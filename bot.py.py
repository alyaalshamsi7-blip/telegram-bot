import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    filters,
)

TOKEN = "8539216702:AAHOoZ8V5_oGsrnhxjg2xg4-4e9kLB7nSDo"

LANG, NAME, STUDENT_ID, PARENT_EMAIL, REQUEST_TYPE, ISSUE = range(6)


def save_to_excel(data):
    file_name = "complaints.xlsx"

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Requests"
        ws.append([
            "Date",
            "Language",
            "Student Name",
            "Student ID",
            "Parent Email",
            "Request Type",
            "Issue Description",
            "Status",
        ])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb["Requests"]
    ws.append([
        data["date"],
        data["lang"],
        data["name"],
        data["student_id"],
        data["parent_email"],
        data["request_type"],
        data["issue"],
        "Pending",
    ])
    wb.save(file_name)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [["عربي", "English"]]
    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        resize_keyboard=True,
        one_time_keyboard=True
    )

    await update.message.reply_text(
        "اختر اللغة / Choose language:",
        reply_markup=reply_markup
    )
    return LANG


async def choose_lang(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_lang = update.message.text.strip()

    if user_lang not in ["عربي", "English"]:
        keyboard = [["عربي", "English"]]
        reply_markup = ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await update.message.reply_text(
            "يرجى اختيار اللغة من الأزرار / Please choose a language from the buttons.",
            reply_markup=reply_markup
        )
        return LANG

    context.user_data["lang"] = user_lang

    if user_lang == "English":
        await update.message.reply_text(
            "Enter student name:",
            reply_markup=ReplyKeyboardRemove()
        )
    else:
        await update.message.reply_text(
            "أدخل اسم الطالب:",
            reply_markup=ReplyKeyboardRemove()
        )

    return NAME


async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()

    if not name:
        if context.user_data["lang"] == "English":
            await update.message.reply_text("Student name is required. Enter student name:")
        else:
            await update.message.reply_text("اسم الطالب مطلوب. أدخل اسم الطالب:")
        return NAME

    context.user_data["name"] = name

    if context.user_data["lang"] == "English":
        await update.message.reply_text("Enter student ID:")
    else:
        await update.message.reply_text("أدخل رقم الطالب:")

    return STUDENT_ID


async def get_student_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    student_id = update.message.text.strip()

    if not student_id:
        if context.user_data["lang"] == "English":
            await update.message.reply_text("Student ID is required. Enter student ID:")
        else:
            await update.message.reply_text("رقم الطالب مطلوب. أدخل رقم الطالب:")
        return STUDENT_ID

    context.user_data["student_id"] = student_id

    if context.user_data["lang"] == "English":
        await update.message.reply_text("Enter parent email address:")
    else:
        await update.message.reply_text("أدخل البريد الإلكتروني لولي الأمر:")

    return PARENT_EMAIL


async def get_parent_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parent_email = update.message.text.strip()

    if not parent_email or "@" not in parent_email or "." not in parent_email:
        if context.user_data["lang"] == "English":
            await update.message.reply_text("Please enter a valid parent email address:")
        else:
            await update.message.reply_text("يرجى إدخال بريد إلكتروني صحيح لولي الأمر:")
        return PARENT_EMAIL

    context.user_data["parent_email"] = parent_email

    if context.user_data["lang"] == "English":
        keyboard = [
            ["Ask about grade calculation"],
            ["Ask about grade distribution"],
            ["Technical issue viewing grades"],
            ["Ask about evaluation method"],
            ["Concern about academic level"],
            ["Remark on results"],
        ]
        reply_markup = ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await update.message.reply_text(
            "Choose request type:",
            reply_markup=reply_markup
        )
    else:
        keyboard = [
            ["الاستفسار عن آلية احتساب الدرجات"],
            ["الاستفسار عن توزيع الدرجات الفرعية"],
            ["مشكلة تقنية في الاطلاع على الدرجات"],
            ["الاستفسار عن آلية التقييم"],
            ["القلق على المستوى الأكاديمي"],
            ["لدي ملاحظة على النتائج"],
        ]
        reply_markup = ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await update.message.reply_text(
            "اختر نوع الطلب:",
            reply_markup=reply_markup
        )

    return REQUEST_TYPE


async def get_request_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    request_type = update.message.text.strip()
    lang = context.user_data["lang"]

    arabic_options = [
        "الاستفسار عن آلية احتساب الدرجات",
        "الاستفسار عن توزيع الدرجات الفرعية",
        "مشكلة تقنية في الاطلاع على الدرجات",
        "الاستفسار عن آلية التقييم",
        "القلق على المستوى الأكاديمي",
        "لدي ملاحظة على النتائج",
    ]

    english_options = [
        "Ask about grade calculation",
        "Ask about grade distribution",
        "Technical issue viewing grades",
        "Ask about evaluation method",
        "Concern about academic level",
        "Remark on results",
    ]

    valid_options = arabic_options if lang == "عربي" else english_options

    if request_type not in valid_options:
        reply_markup = ReplyKeyboardMarkup(
            [[opt] for opt in valid_options],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        if lang == "English":
            await update.message.reply_text(
                "Please choose a request type from the buttons:",
                reply_markup=reply_markup
            )
        else:
            await update.message.reply_text(
                "يرجى اختيار نوع الطلب من الأزرار:",
                reply_markup=reply_markup
            )
        return REQUEST_TYPE

    context.user_data["request_type"] = request_type

    if lang == "English":
        await update.message.reply_text(
            "Write the issue description:",
            reply_markup=ReplyKeyboardRemove()
        )
    else:
        await update.message.reply_text(
            "اكتب وصف المشكلة أو تفاصيل الاستفسار:",
            reply_markup=ReplyKeyboardRemove()
        )

    return ISSUE


async def get_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    issue = update.message.text.strip()

    if not issue:
        if context.user_data["lang"] == "English":
            await update.message.reply_text("Issue description is required. Please write the issue:")
        else:
            await update.message.reply_text("وصف المشكلة مطلوب. يرجى كتابة وصف المشكلة:")
        return ISSUE

    context.user_data["issue"] = issue
    context.user_data["date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    save_to_excel(context.user_data)

    if context.user_data["lang"] == "English":
        await update.message.reply_text(
            "Your request has been received successfully.\n"
            "The parent email is required so the school can contact you.\n"
            "Your data has been saved.\n"
            "Status: Pending",
            reply_markup=ReplyKeyboardRemove()
        )
    else:
        await update.message.reply_text(
            "تم استلام طلبك بنجاح.\n"
            "البريد الإلكتروني لولي الأمر مطلوب ليصل إليه رد المدرسة.\n"
            "تم حفظ البيانات.\n"
            "الحالة: Pending",
            reply_markup=ReplyKeyboardRemove()
        )

    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("lang") == "English":
        await update.message.reply_text(
            "Operation cancelled.",
            reply_markup=ReplyKeyboardRemove()
        )
    else:
        await update.message.reply_text(
            "تم إلغاء العملية.",
            reply_markup=ReplyKeyboardRemove()
        )

    context.user_data.clear()
    return ConversationHandler.END


def main():
    app = Application.builder().token(TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            LANG: [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_lang)],
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_student_id)],
            PARENT_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_parent_email)],
            REQUEST_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_request_type)],
            ISSUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_issue)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv_handler)
    app.run_polling()


if __name__ == "__main__":
    main()